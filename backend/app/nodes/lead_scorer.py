"""
Node 5: Lead Scorer
====================
Scores each lead 0-100 based on signals that indicate sales-readiness:
- Google rating & review count (legitimacy)
- Email quality (owner email > generic)
- Website presence (digital maturity)
- Review velocity (active business)

Node 6: Delivery
=================
Writes the final dataset to CSV and XLSX.
"""

import os
import csv
import logging
from datetime import datetime
from pathlib import Path

from app.agents.graph import LeadState

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "/tmp/leads"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ── Scorer ────────────────────────────────────────────────────────────────────

def _score_lead(lead: dict) -> int:
    score = 0

    # Rating signal (max 25 pts)
    rating = float(lead.get("rating", 0) or 0)
    if rating >= 4.5:  score += 25
    elif rating >= 4.0: score += 18
    elif rating >= 3.0: score += 10

    # Review count signal (max 20 pts)
    reviews = int(lead.get("review_count", 0) or 0)
    if reviews >= 100:  score += 20
    elif reviews >= 50: score += 15
    elif reviews >= 20: score += 10
    elif reviews >= 5:  score += 5

    # Email quality (max 30 pts)
    if lead.get("email_verified"):    score += 30
    elif lead.get("email_catchall"):  score += 20
    elif lead.get("email_valid"):     score += 10

    # Owner vs generic (max 15 pts)
    owner_kws = ["owner", "ceo", "founder", "president", "manager", "director"]
    position  = (lead.get("owner_position") or "").lower()
    email_pfx = (lead.get("primary_email") or "").split("@")[0].lower()
    if any(kw in position for kw in owner_kws):   score += 15
    elif any(kw in email_pfx for kw in owner_kws): score += 10

    # Website presence (max 10 pts)
    if lead.get("website"): score += 10

    return min(100, score)


def score_leads_node(state: LeadState) -> LeadState:
    leads = state.get("verified_leads", [])
    logger.info(f"[scorer] Scoring {len(leads)} leads")

    scored = []
    for lead in leads:
        lead = {**lead, "score": _score_lead(lead)}
        scored.append(lead)

    # Sort by score descending
    scored.sort(key=lambda l: l["score"], reverse=True)

    logger.info(
        f"[scorer] Score distribution: "
        f"high(≥70)={sum(1 for l in scored if l['score']>=70)}, "
        f"mid(40-69)={sum(1 for l in scored if 40<=l['score']<70)}, "
        f"low(<40)={sum(1 for l in scored if l['score']<40)}"
    )

    return {**state, "scored_leads": scored}


# ── Delivery ──────────────────────────────────────────────────────────────────

COLUMNS = [
    "score", "name", "primary_email", "email_verified", "email_catchall",
    "email_status", "owner_name", "owner_position", "phone", "website",
    "address", "rating", "review_count", "category",
    "latitude", "longitude", "source",
]


def deliver_leads_node(state: LeadState) -> LeadState:
    leads    = state.get("scored_leads", [])
    query    = state.get("raw_query", "leads")[:40].replace(" ", "_").replace("/", "-")
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{query}_{ts}"

    # ── CSV ──────────────────────────────────────────────────────────────────
    csv_path = OUTPUT_DIR / f"{filename}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(leads)

    # ── XLSX (if openpyxl available) ─────────────────────────────────────────
    xlsx_path = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows with conditional score coloring
        green_fill  = PatternFill("solid", fgColor="C6EFCE")  # score ≥ 70
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")  # score 40-69
        red_fill    = PatternFill("solid", fgColor="FFC7CE")  # score < 40

        for row_idx, lead in enumerate(leads, 2):
            score = lead.get("score", 0)
            row_fill = green_fill if score >= 70 else (yellow_fill if score >= 40 else red_fill)

            for col_idx, col_name in enumerate(COLUMNS, 1):
                val  = lead.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill

        # Auto-width columns
        for col_idx, col_name in enumerate(COLUMNS, 1):
            max_len = max(
                len(str(col_name)),
                *(len(str(lead.get(col_name, ""))) for lead in leads[:50])
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        xlsx_path = OUTPUT_DIR / f"{filename}.xlsx"
        wb.save(xlsx_path)
        logger.info(f"[deliver] XLSX saved: {xlsx_path}")

    except ImportError:
        logger.warning("[deliver] openpyxl not installed — skipping XLSX export")
    except Exception as e:
        logger.error(f"[deliver] XLSX error: {e}")

    logger.info(f"[deliver] CSV saved: {csv_path} ({len(leads)} leads)")

    return {
        **state,
        "final_csv_path": str(csv_path),
        "final_xlsx_path": str(xlsx_path) if xlsx_path else "",
        "status": "done",
    }